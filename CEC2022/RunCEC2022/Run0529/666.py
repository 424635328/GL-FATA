import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. 原始 Markdown 表格数据（使用 r"""...""" 声明为原始字符串，防止 \t 被解析为制表符）
markdown_table = r"""
| Function | Metric | GL-FATA | MFATA-Levy | IMFATA | ASFSSA | PSO | FATA | GWO | SSA |
| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |
| **F1 (Unimodal)** | Min | $3.00 \times 10^2$ | $3.01 \times 10^2$ | $3.01 \times 10^2$ | $3.00 \times 10^2$ | $3.00 \times 10^2$ | $3.00 \times 10^2$ | $1.58 \times 10^3$ | $3.00 \times 10^2$ |
| | Mean | $3.00 \times 10^2$ | $5.05 \times 10^2$ | $3.39 \times 10^2$ | $3.00 \times 10^2$ | $5.64 \times 10^2$ | $4.12 \times 10^2$ | $8.29 \times 10^3$ | $3.00 \times 10^2$ |
| | Std | $2.69 \times 10^{-12}$ | $4.10 \times 10^2$ | $5.34 \times 10^1$ | $1.72 \times 10^{-13}$ | $1.45 \times 10^3$ | $3.44 \times 10^2$ | $3.64 \times 10^3$ | $1.23 \times 10^{-13}$ |
| **F6 (Multimodal)** | Mean | $4.89 \times 10^3$ | $2.20 \times 10^4$ | $5.29 \times 10^4$ | $8.30 \times 10^3$ | $5.53 \times 10^4$ | $4.40 \times 10^3$ | $2.48 \times 10^5$ | $9.90 \times 10^3$ |
| | Std | $5.25 \times 10^3$ | $1.45 \times 10^4$ | $3.35 \times 10^4$ | $7.55 \times 10^3$ | $2.71 \times 10^5$ | $1.72 \times 10^3$ | $8.84 \times 10^5$ | $7.44 \times 10^3$ |
| **F9 (Composition)** | Mean | $2.48 \times 10^3$ | $2.48 \times 10^3$ | $2.48 \times 10^3$ | $2.48 \times 10^3$ | $2.51 \times 10^3$ | $2.48 \times 10^3$ | $2.50 \times 10^3$ | $2.48 \times 10^3$ |
| | Std | $1.35 \times 10^{-2}$ | $1.54 \times 10^{-1}$ | $2.04 \times 10^{-1}$ | $7.74 \times 10^{-12}$ | $3.28 \times 10^1$ | $1.10 \times 10^0$ | $2.04 \times 10^1$ | $2.17 \times 10^{-2}$ |
| **F12 (Composition)** | Mean | $2.96 \times 10^3$ | $2.95 \times 10^3$ | $2.95 \times 10^3$ | $2.98 \times 10^3$ | $3.01 \times 10^3$ | $2.99 \times 10^3$ | $2.97 \times 10^3$ | $3.01 \times 10^3$ |
| | Std | $1.21 \times 10^1$ | $1.12 \times 10^1$ | $1.27 \times 10^1$ | $2.82 \times 10^1$ | $5.16 \times 10^1$ | $2.89 \times 10^1$ | $2.28 \times 10^1$ | $5.27 \times 10^1$ |
"""

# 2. 解析 LaTeX 科学计数法格式为 Python 浮点数
def parse_latex_scientific(val):
    val = val.strip().replace("**", "")
    
    # 核心容错处理：将可能已经被解释为制表符的 \t 还原为 \t 字符
    val = val.replace('\t', '\\t')
    
    # 正则表达式：兼容 literal \times 和已被转义的 \\times
    match = re.search(r'\$?([-+]?[0-9]*\.?[0-9]+)\s*\\times\s*10\^\{?(-?[0-9]+)\}?\$?', val)
    if match:
        base = float(match.group(1))
        exponent = int(match.group(2))
        return base * (10 ** exponent)
    return val

# 3. 解析 Markdown 表格文本
lines = [line.strip() for line in markdown_table.strip().split('\n') if line.strip()]
headers = []
rows = []

for idx, line in enumerate(lines):
    if not line.startswith('|'):
        continue
    # 拆分并去除多余的首尾空元素
    parts = [p.strip() for p in line.split('|')[1:-1]]
    if idx == 0:
        headers = parts
    elif idx == 1:
        # 跳过 Markdown 分割线（如 | :--- | :--- |）
        continue
    else:
        processed_row = []
        for col_idx, part in enumerate(parts):
            if col_idx >= 2:  # 算法寻优结果数值列
                processed_row.append(parse_latex_scientific(part))
            else:  # Function 和 Metric 列
                processed_row.append(part.replace("**", ""))
        rows.append(processed_row)

# 4. 创建 Excel 工作簿并应用排版样式
wb = Workbook()
ws = wb.active
ws.title = "寻优结果统计"

# 启用网格线可见
ws.views.sheetView[0].showGridLines = True

# 写入大标题
ws.merge_cells("A1:J1")
ws["A1"] = "表3.3 CEC 2022部分代表性函数的寻优结果统计"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# 写入表头列名并美化
start_row = 3  # 从第3行开始写入表格主体
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col_idx, value=header)
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ws.row_dimensions[start_row].height = 25

# 定义淡灰色细边框样式
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# 写入数据行
for row_idx, row in enumerate(rows, start_row + 1):
    ws.row_dimensions[row_idx].height = 20
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        
        # 格式化数值列为科学计数法
        if col_idx > 2:
            cell.number_format = '0.00E+00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

# 5. 自动合并 Function 列中重复的行
start_merge_row = start_row + 1
current_func = ""

for row_idx in range(start_row + 1, start_row + len(rows) + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val != "":
        # 如果不是第一组，且上一组有多个行，则进行合并
        if current_func != "" and row_idx - 1 > start_merge_row:
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=row_idx-1, end_column=1)
            # 重新应用合并后左上角单元格的样式（居中、粗体）
            merged_cell = ws.cell(row=start_merge_row, column=1)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            merged_cell.font = Font(name="Calibri", size=10, bold=True)
        current_func = val
        start_merge_row = row_idx

# 合并最后一组（如 F12）
if start_merge_row < start_row + len(rows):
    ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=start_row + len(rows), end_column=1)
    merged_cell = ws.cell(row=start_merge_row, column=1)
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    merged_cell.font = Font(name="Calibri", size=10, bold=True)

# 6. 自动调整列宽以适应内容
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        # 跳过标题行，避免标题过长导致列宽失真
        if cell.row == 1:
            continue
        # 将科学计数法的显示格式也估算进列宽中
        if isinstance(cell.value, float):
            val_str = f"{cell.value:.2E}"
        else:
            val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    # 适当留白
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# 7. 保存文件
output_filename = "CEC2022_Optimization_Results.xlsx"
wb.save(output_filename)
print(f"Excel文件已成功导出并保存为: {output_filename}")